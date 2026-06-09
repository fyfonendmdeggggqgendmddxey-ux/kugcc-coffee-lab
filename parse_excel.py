import openpyxl
import json

wb = openpyxl.load_workbook("/Users/norokazuhito/Downloads/Coffee_Grinders_Micron_Master_v2.xlsx", data_only=True)
result = {}

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    description = sheet.cell(row=2, column=1).value
    
    header_row_idx = -1
    headers = []
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if any(x and "粒径 (µm)" in str(x) for x in row):
            header_row_idx = row_idx
            headers = [str(x) if x is not None else f"col_{i}" for i, x in enumerate(row)]
            break
            
    if header_row_idx == -1:
        continue
        
    micron_col = next((i for i, h in enumerate(headers) if "粒径" in h), -1)
    
    table = []
    for row in sheet.iter_rows(min_row=header_row_idx+1, values_only=True):
        if micron_col == -1 or row[micron_col] is None:
            continue
            
        microns = float(row[micron_col])
        label_parts = []
        for i, h in enumerate(headers):
            val = row[i]
            if val is None: continue
            
            if "ギア" in h: label_parts.append(f"{val}ギア")
            elif "グリッド" in h: label_parts.append(f"{val}グリッド")
            elif "目盛り" in h: label_parts.append(str(val))
            elif "総クリック" in h:
                if sheet_name not in ["Lagom casa", "Mahlkönig EK43", "Lagom P64", "Fellow Ode Gen 2"]:
                    label_parts.append(f"{val}")
        
        label = " ".join(label_parts).strip()
        if not label:
            clicks_col = next((i for i, h in enumerate(headers) if "総クリック" in h), -1)
            if clicks_col != -1 and row[clicks_col] is not None:
                label = str(row[clicks_col])
                
        table.append({"label": label, "microns": microns})
        
    result[sheet_name] = {
        "description": str(description),
        "table": table
    }

with open("utils/unified_grinders.json", "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print("Done")
