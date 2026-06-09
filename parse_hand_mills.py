import openpyxl
import json
import os

file_path = "/Users/norokazuhito/Downloads/Coffee_Grinder_Master_Database.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

sheets_to_parse = ['Timemoreシリーズ', 'Comandanteシリーズ', 'その他高性能ミル']

result = {}

# S3 overrides based on user input
s3_overrides = {
    "6.5": 842,
    "8.0": 960
}

for sheet_name in sheets_to_parse:
    if sheet_name not in wb.sheetnames:
        continue
    sheet = wb[sheet_name]
    
    # Grinder names are in row 4
    # Starting from col 2 (B)
    for col_idx in range(2, sheet.max_column + 1, 4):
        grinder_name = sheet.cell(row=4, column=col_idx).value
        if not grinder_name:
            continue
            
        data_points = []
        for row_idx in range(6, sheet.max_row + 1):
            setting = sheet.cell(row=row_idx, column=col_idx).value
            micron = sheet.cell(row=row_idx, column=col_idx + 1).value
            
            if setting is not None and micron is not None:
                # Store setting as string to be consistent with electric mills
                data_points.append({"label": str(setting), "microns": float(micron)})
                
        if data_points:
            # Apply overrides for Timemore S3
            if grinder_name == "Timemore S3":
                # Override existing 8.0
                for dp in data_points:
                    if dp["label"] == "8.0":
                        dp["microns"] = s3_overrides["8.0"]
                # Insert 6.5 and sort
                data_points.append({"label": "6.5", "microns": s3_overrides["6.5"]})
                
                # Sort numerically
                def parse_num(s):
                    import re
                    m = re.search(r"[\d.]+", s)
                    return float(m.group(0)) if m else 0.0
                data_points.sort(key=lambda x: parse_num(x["label"]))

            result[grinder_name] = {
                "description": f"Excelデータからインポート ({sheet_name})",
                "table": data_points
            }

output_path = os.path.join(os.path.dirname(__file__), "utils", "hand_grinders.json")
with open(output_path, "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f"Parsed {len(result)} hand mills.")
